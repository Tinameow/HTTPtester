# -*- coding: utf-8 -*-

import xlrd
import openpyxl
import os
from datetime import datetime
from collections import OrderedDict

WORKBOOK = 'data/projectmanager.xlsx'
DATADIR = 'data'



def exist_proj_id(proj_id:str) -> bool:
    '''return whether the proj id exists in db'''
    return proj_id + '.xlsx' in os.listdir(DATADIR)



def exist_tid(tid:str) -> bool:
    '''return whether the tid exists in db'''
    return tid + '.xlsx' in os.listdir(DATADIR)



def exist_cid(cid:int, workbook_name:str) -> bool:
    '''return whether the cid exists in db; the workbookname is either proj_id or tid'''
    return 1 <= cid <= xlrd.open_workbook(os.path.join(DATADIR,workbook_name+'.xlsx')).sheet_by_name('testcases').nrows



def is_valid_case(name, method, url,req_header, data, exp_status_code, assert_key, exp_value)->bool:
    '''return whether the case's format is valid'''
    return (name and method and url) and ((assert_key and exp_value) or (not assert_key and not exp_value)) \
    and (method == 'POST' or (method != 'POST' and not data))



def get_projects() -> [dict]:
    '''return all the projects' infomation'''
    data = []
    with xlrd.open_workbook(WORKBOOK) as workbook:
        sheet = workbook.sheet_by_index(0)
        for i in range(1, sheet.nrows):
            data.append(OrderedDict({"proj_name":sheet.cell(i,1).value, 
                                     "proj_id":sheet.cell(i,0).value,
                                     "proj_info":sheet.cell(i,2).value}))
    return data

   

def add_proj(name:str, info:str) -> str:
    '''add a project by adding it to the proj file and creating its proj workbook'''
    workbook = openpyxl.load_workbook(WORKBOOK)
    sheet = workbook['Sheet']
    # calculate the new proj id 
    if sheet.max_row == 1:
        new_id = 'A00001'
    else:
        new_id = 'A' + (str(int((sheet.cell(row = sheet.max_row, column = 1).value)[1:])+1)).zfill(5)
    sheet.append([new_id, name, info])
    workbook.save(filename = WORKBOOK)
    # add new proj workbook to save testcases
    newbook = openpyxl.Workbook()
    newbook['Sheet'].title = 'testcases'
    newbook['testcases'].append(['name','method','url','request_header',
                             'data','exp_status_code','assert_key','exp_value'])
    newbook.save(filename = os.path.join(DATADIR, new_id + '.xlsx'))
    return new_id



def delete_proj(proj_id:str):
    '''delete the project by deleting it in the proj file and its proj workbook'''
    workbook = openpyxl.load_workbook(WORKBOOK)
    sheet = workbook['Sheet']
    for i,row in enumerate(sheet.rows,1):
        if row[0].value == proj_id:
            sheet.delete_rows(i)
            workbook.save(filename = WORKBOOK)
            break 
    for filename in os.listdir(DATADIR):
        if proj_id in filename:
            os.remove(os.path.join(DATADIR,filename))
    return None



def get_testcases(workbook_name:str, key:str = None, cids:[int] = None, form:str = 'dict') -> [list]:
    '''return the testcases in the project's workbook and '''
    data = []
    with xlrd.open_workbook(os.path.join(DATADIR, workbook_name+'.xlsx')) as workbook:
        sheet = workbook.sheet_by_name('testcases')
        if key:
            for i in range(1, sheet.nrows):
                if key in sheet.cell(i,0).value:
                    data.append(_get_testcase(sheet, i)) if form == 'dict' else data.append(sheet.row_values(int(i)))
        elif cids:
             for i in cids:
                if 1 <= int(i) < sheet.nrows:
                    data.append(_get_testcase(sheet, i)) if form == 'dict' else data.append(sheet.row_values(int(i)))
        else:
            for i in range(1, sheet.nrows):
                data.append(_get_testcase(sheet, i)) if form == 'dict' else data.append(sheet.row_values(int(i)))

    return data



def add_testcases(workbook_name: str, testcases: [list]): 
    '''add testcases to the workbook'''
    workbook = openpyxl.load_workbook(os.path.join(DATADIR,workbook_name+'.xlsx'))
    sheet = workbook['testcases']
    for testcase in testcases:
        print(testcase)
        sheet.append(testcase)    
    workbook.save(os.path.join(DATADIR, workbook_name+'.xlsx'))
    



def delete_cases(proj_id, cids:[int]):
    '''delete cases in the proj testcases'''
    workbook = openpyxl.load_workbook(os.path.join(DATADIR,proj_id+'.xlsx'))
    sheet = workbook['testcases']
    for i in reversed(cids):
        if 1 <= int(i) < sheet.max_row:
            sheet.delete_rows(int(i)+1)
    workbook.save(filename = os.path.join(DATADIR,proj_id+'.xlsx'))       



def create_test(proj_id,cids:[int]) -> str:
    '''creat a test with the proj id and cids'''
    # get the current time and generate the tid
    start_time = datetime.now()
    tid = _generate_tid(proj_id,start_time)
    # create a new workbook
    workbook = openpyxl.Workbook()
    # add basic information 
    testinfo = workbook['Sheet']
    testinfo.append(['tid','status','total_cases','tested_cases','passed_cases','start_time','end_time'])
    testcases = workbook.create_sheet(title = 'testcases')
    testcases.append(['name','method','url','request_header',
                  'data','exp_status_code','assert_key','exp_value','act_status_code','act_value','act_response','result'])    
    # add testcases 
    cases = get_testcases(proj_id, cids = cids, form = 'list')
    testinfo.append([tid,'Processing',len(cids) if cids else len(cases),0,0,start_time.strftime('%Y-%m-%d %H:%M:%S')])
    for testcase in cases:
        testcases.append(testcase+['','','',''])    

    # update the total number
    testinfo.cell(2,3).value = testcases.max_row - 1
    workbook.save(os.path.join(DATADIR, tid+'.xlsx'))

    return tid 



def update_test_result(tid:str, cid: int, result: tuple):
    '''update the test result with the result'''
    workbook = openpyxl.load_workbook(os.path.join(DATADIR,tid+'.xlsx'))
    sheet = workbook['testcases']

    sheet.cell(cid+1, 9, value = result[0])
    sheet.cell(cid+1, 10, value = result[1])
    sheet.cell(cid+1, 11, value = result[2])
    sheet.cell(cid+1, 12, value = result[3])
    workbook['Sheet'].cell(2,4).value += 1
    if result[3] == 'PASS':
        workbook['Sheet'].cell(2,5).value += 1  
    if workbook['Sheet'].cell(2,4).value == workbook['Sheet'].cell(2,3).value:
        workbook['Sheet'].cell(2,2).value = 'Done'
        workbook['Sheet'].cell(2,7).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    workbook.save(os.path.join(DATADIR,tid+'.xlsx'))



def get_tests(proj_id) -> [dict]:
    '''return all the tests' basic info for the proj'''
    data = []
    for filename in os.listdir(DATADIR):
        if proj_id in filename and filename != proj_id+'.xlsx':
            data.append(get_test(filename.split('.')[0]))
    return data



def get_test(tid) -> dict:
    '''return single test's basic info'''
    with xlrd.open_workbook(os.path.join(DATADIR,tid+'.xlsx')) as workbook:
        sheet = workbook.sheet_by_name('Sheet')
        return OrderedDict({'tid': sheet.cell(1,0).value,
                            'status': sheet.cell(1,1).value,
                            'total_cases': int(sheet.cell(1,2).value),
                            'tested_cases': int(sheet.cell(1,3).value),
                            'passed_cases': int(sheet.cell(1,4).value),
                            'start_time': sheet.cell(1,5).value,
                            'end_time': sheet.cell(1,6).value} ) 



def _get_testcase(sheet, cid: int) -> dict:
    if sheet.ncols <= 8:
        return OrderedDict({"cid":cid,
                    "Name":sheet.cell(cid,0).value, 
                    "Method":sheet.cell(cid,1).value,
                    "URL":sheet.cell(cid,2).value,
                    "Request_header":sheet.cell(cid,3).value,
                    "data":sheet.cell(cid,4).value,
                    "exp_status_code":int(sheet.cell(cid,5).value) if sheet.cell(cid,5).value.isdigit() else '',
                    "assert_key":sheet.cell(cid,6).value,
                    "exp_value":sheet.cell(cid,7).value})

    return OrderedDict({"cid":cid,
                        "Name":sheet.cell(cid,0).value, 
                        "Method":sheet.cell(cid,1).value,
                        "URL":sheet.cell(cid,2).value,
                        "Request_header":sheet.cell(cid,3).value,
                        "data":sheet.cell(cid,4).value,
                        "exp_status_code":int(sheet.cell(cid,5).value) if sheet.cell(cid,5).value.isdigit() else '',
                        "assert_key":sheet.cell(cid,6).value,
                        "exp_value":sheet.cell(cid,7).value,
                        'act_status_code':sheet.cell(cid,8).value,
                        'act_value':sheet.cell(cid,9).value,
                        'act_response':sheet.cell(cid,10).value,
                        'result':sheet.cell(cid,11).value})


def _generate_tid(proj_id, start_time):
    return proj_id + '_'+ start_time.strftime('%Y%m%d%H%M%S')